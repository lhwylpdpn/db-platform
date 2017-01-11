#!/usr/bin/python
# -*- coding: UTF-8 -*- 



import xlrd
from openpyxl import Workbook
from openpyxl import load_workbook
import pymysql
import os
import sys
import time
import csv
import sys
import datetime
reload(sys)
sys.setdefaultencoding('utf-8')
def import_excel():
	test = []
	test= os.listdir("src")

	newlist=[]
	statinfo=[]
	for names in test:

		if names.endswith(".csv"):
			#statinfo.append(os.stat(os.getcwd()+"/src/"+names).st_ctime)
			newlist.append(names)
	for r in newlist:
		print(r)
	
		if  not r.find("Daily_lchy"):#投放转化
			filename=os.getcwd()+"/src/"+r
			filenode=open(filename)
			reader=csv.reader(filenode)
			sql="truncate zilong_report.retention;insert into zilong_report.retention values "
			for row in reader:
				sql=sql+"('"+"','".join(row)+"'),"
			sql=sql.strip(',')+";"
			print("1sql ok")
				#time.sleep(10)
			cur_1.execute(sql)
	
		if  not r.find("Daily_tfzh"):#投放转化
			filename=os.getcwd()+"/src/"+r
			filenode=open(filename)
			reader=csv.reader(filenode)
			sql="truncate zilong_report.ad_action;insert into zilong_report.ad_action values "
			for row in reader:
				sql=sql+"('"+"','".join(row)+"'),"
			sql=sql.strip(',')+";"
			print("2sql ok")
				#time.sleep(10)
			cur_1.execute(sql)

		if  not r.find("Daily_zbhs"):#投放转化
			filename=os.getcwd()+"/src/"+r
			filenode=open(filename)
			reader=csv.reader(filenode)
			sql="truncate zilong_report.re_money;insert into zilong_report.re_money values "
			for row in reader:
				sql=sql+"('"+"','".join(row)+"'),"
			sql=sql.strip(',')+";"
			print("3sql ok")
				#time.sleep(10)
			cur_1.execute(sql)

		if  not r.find("LaunchPaymentAll"):#投放转化
			filename=os.getcwd()+"/src/"+r
			filenode=open(filename)
			reader=csv.reader(filenode)
			sql="truncate zilong_report.spend;insert into zilong_report.spend values "
			for row in reader:
				sql=sql+"('"+"','".join(row)+"'),"
			sql=sql.strip(',')+";"
			print("4sql ok ")
			cur_1.execute(sql)
def export():
	sql="DROP TABLE IF EXISTS temp1;CREATE TABLE temp1 AS SELECT * FROM (SELECT DATE_FORMAT(b.DATE,'%Y-%m-%d') AS DATE,a.staff,b.channel_name,b.agent FROM `ad_action` b LEFT JOIN spend a ON a.date=b.date AND a.channel_name=b.channel_name AND a.agent=b.agent   UNION  SELECT DATE_FORMAT(a.DATE,'%Y-%m-%d') AS DATE,a.staff,a.channel_name,a.agent FROM `ad_action` b RIGHT JOIN spend a ON a.date=b.date AND a.channel_name=b.channel_name AND a.agent=b.agent  ) a order by date,channel_name,agent;delete from temp1 where date is null"
	cur_1.execute(sql)
	sql="SELECT a.date,a.staff,a.channel_name, a.agent,a.ad_click,a.ad_action,a.ad_action_new,a.ad_account_new,a.ad_account_new_pay,a.ad_account_new_paymoney, a.fufeilv,ROUND(b.dis_spend/a.ad_account_new,2) AS cpa,b.dis_spend FROM (SELECT b.date,b.staff,b.channel_name, b.agent,a.ad_click,a.ad_action,a.ad_action_new,a.ad_account_new,a.ad_account_new_pay,a.ad_account_new_paymoney ,CONCAT(ROUND(a.ad_account_new_pay/a.ad_account_new*100,2),'%') AS fufeilv FROM ad_action a RIGHT JOIN temp1 b ON a.date=b.date AND a.channel_name=b.channel_name AND a.agent=b.agent ) a left join (SELECT DATE ,channel_name,agent,SUM(dis_spend) AS dis_spend FROM spend GROUP BY DATE,channel_name,agent ) b on a.date=b.date AND a.channel_name=b.channel_name AND a.agent=b.agent ORDER BY a.date,a.channel_name,a.agent"
	cur_1.execute(sql)
	res=cur_1.fetchall()
	book=load_workbook('test.xlsx')
	export_sheet=book.get_sheet_by_name("test")
	if len(res)>1:

		i=4
		# export_sheet.cell(row = 1, column = 2, value = "日期")
		# export_sheet.cell(row = 1, column = 3, value = "负责人")
		# export_sheet.cell(row = 1, column = 4, value = "渠道名称")
		# export_sheet.cell(row = 1, column = 5, value = "代理公司")
		# export_sheet.cell(row = 1, column = 6, value = "广告点击设备")
		# export_sheet.cell(row = 1, column = 7, value = "广告激活设备")
		# export_sheet.cell(row = 1, column = 8, value = "广告新增设备")
		# export_sheet.cell(row = 1, column = 9, value = "广告新增账号")
		# export_sheet.cell(row = 1, column = 10, value ="广告新增付费账号")
		# export_sheet.cell(row = 1, column = 11, value ="广告新增付费金额")
		# export_sheet.cell(row = 1, column = 12, value ="付费率")
		# export_sheet.cell(row = 1, column = 13, value ="CPA")
		# export_sheet.cell(row = 1, column = 14, value ="花费金额")

		for r in res:
			export_sheet.cell(row = i, column = 2, value = r[0])
			export_sheet.cell(row = i, column = 3, value = r[1])
			export_sheet.cell(row = i, column = 4, value = r[2])
			export_sheet.cell(row = i, column = 5, value = r[3])
			export_sheet.cell(row = i, column = 6, value = r[4])
			export_sheet.cell(row = i, column = 7, value = r[5])
			export_sheet.cell(row = i, column = 8, value = r[6])
			export_sheet.cell(row = i, column = 9, value = r[7])
			export_sheet.cell(row = i, column = 10, value = r[8])
			export_sheet.cell(row = i, column = 11, value = r[9])
			export_sheet.cell(row = i, column = 12, value = r[10])
			export_sheet.cell(row = i, column = 13, value = r[11])
			export_sheet.cell(row = i, column = 14, value = r[12])
			export_sheet.cell(row = i, column = 14, value = r[12])
			i=i+1
		book.save('test.xlsx')
		print("1ok")
	print time.asctime(time.localtime(time.time()))
# part2 收入
	sql="SELECT a.*,0.686 FROM (SELECT b.date AS data1 ,b.staff,b.`channel_name` AS c1,b.`agent` AS a1,a.* FROM `re_money` a RIGHT JOIN temp1 b ON a.`date`=b.date AND a.`channel_name`=b.channel_name AND a.`agent`=b.agent ORDER BY b.date,b.channel_name,b.agent ) a left join `dis_result` b on a.data1=b.`date` order by a.data1,a.c1,a.a1"
	cur_1.execute(sql)
	res=cur_1.fetchall()
	if len(res)>1:

		for j in xrange(0,133):
			export_sheet.cell(row = 1, column = 16+j, value = "+"+str(j)+"收入")
			i=4
			for r in res:
				
				export_sheet.cell(row = i, column = 16+j, value = r[8+j])
				i=i+1
		i=4
		for r in res:
			#print(r[141],r[67],r[67]*r[141],r[140]*r[141])
		
			export_sheet.cell(row = i, column = 149, value = r[68])
			export_sheet.cell(row = i, column = 150, value = max(r[9:141]))
			if  r[68] is not None:
				
				export_sheet.cell(row = i, column = 151, value = float(r[68])*float(r[373]))
			if  max(r[9:141]) is not None:
				export_sheet.cell(row = i, column = 152, value = float(max(r[9:141]))*float(r[373]))
			i=i+1
# part3 回收计算
	book.save('test.xlsx')	
	print("2ok")
	print time.asctime(time.localtime(time.time()))
	sql="SELECT a.*,b.* FROM `re_money` b RIGHT JOIN  ( SELECT a.*,0.686 FROM (SELECT a.*,b.dis_spend FROM temp1 a LEFT JOIN (SELECT DATE ,channel_name,agent,SUM(dis_spend) AS dis_spend FROM spend GROUP BY DATE,channel_name,agent ) b   ON a.date=b.date AND a.channel_name=b.channel_name AND a.agent=b.agent) a LEFT JOIN dis_result b  ON a.date=b.date ORDER BY a.date) a ON  a.date=b.date AND a.channel_name=b.channel_name AND a.agent=b.agent order by a.date ,a.channel_name ,a.agent"
	cur_1.execute(sql)
	res=cur_1.fetchall()
	if len(res)>1:
		i=3
		for r in res:
			i=i+1
			for j in xrange(154,286):
				if r[5] is not None and r[4] is not None and r[10+j-154] is not None and r[4]!=0:
					export_sheet.cell(row = i, column = j,value=str(round(r[10+j-154]*float(r[5])/r[4]*100,2))+"%")
			if r[215] is not None and r[5] is not None and r[4] is not None and r[4]!=0:
				export_sheet.cell(row = i, column = 287,value=str(round(r[69]*float(r[5])/r[4]*100,2))+"%")
			if max(r[155:287]) is not None and r[5] is not None and r[4] is not None and r[4]!=0:
				export_sheet.cell(row = i, column = 288,value=str(round(max(r[10:141])*float(r[5])/r[4]*100,2))+"%")

#part4 留存
	book.save('test.xlsx')
	print("3ok")
	print time.asctime(time.localtime(time.time()))
	sql="SELECT * FROM `temp1` a LEFT JOIN `Retention` b  ON  a.date=b._date AND a.channel_name=b.channel_name AND a.agent=b.agent ORDER BY a.date,a.channel_name,a.agent"
	cur_1.execute(sql)
	res=cur_1.fetchall()
	if len(res)>1:
		i=3
		for r in res:
			i=i+1
			for j in xrange(290,320):

				if r[9+j-290] is not None and r[9+j-290]!="-":
					try:
						export_sheet.cell(row = i, column = j,value=round(float(r[9+j-290]),2))
					except Exception as e:
						print(e)
				
				if r[9+j-290] is not None and r[9+j-290]!="-" and r[8] is not None and r[8]!="-":
					export_sheet.cell(row = i, column = j+31,value=round(float(r[9+j-290])*float(r[8]),0))





	book.save('test.xlsx')
	print("4ok")
	print time.asctime(time.localtime(time.time()))
def test():
	book=load_workbook('test.xlsx')
	export_sheet=book.get_sheet_by_name("test") 
	for i in range(3,export_sheet.max_row):

		for j in range(1,export_sheet.max_column):
	

			export_sheet.cell(row=i,column=j,value="")
	book.save('test.xlsx')




################ 为了web计算时间所用##############################


time1=datetime.datetime.now()
def test_start():
	file_object = open('status.json','w')
	time1= datetime.datetime.now()#starttime
	file_object.write('{"update_time":"'+str(time1)+'","status":0,"time":"0"}')

	file_object.close()
def test_ok():
	file_object = open('status.json','w')
	time2=datetime.datetime.now()#oktime
	file_object.write('{"update_time":"'+str(time2)+'","status":1,"time":"'+ str((time2 - time1 ).seconds) +'"}')

	file_object.close()
def test_er():
	file_object = open('status.json','w')
	time3= datetime.datetime.now()#errtime
	file_object.write('{"update_time":"'+str(time3)+'","status":2,"time":"'+str((time3 -time1 ).seconds)+'"}')

	file_object.close()


############################################################

if __name__ == '__main__':
	test_start()
	try:
		conn=pymysql.connect(host='localhost',user='root',passwd='123456',db='zilong_report',port=3306)
		cur_1=conn.cursor()
		import_excel()
		export()

		cur_1.close()
		conn.commit()
		conn.close()
		print time.asctime(time.localtime(time.time()))
		test_ok()
	except Exception, e:
		print(e)
		test_er()
	print time.asctime(time.localtime(time.time()))


	# sql="insert into zilong_report.re_money values (truncate zilong_report.re_money;insert into zilong_report.re_money values ('2016-11-06','Aduu','Aduu','6.0','6.0','6.0','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-');insert into zilong_report.re_money values ('2016-11-04','Aduu','Aduu','0.0','6.0','6.0','6.0','6.0','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-');insert into zilong_report.re_money values ('2016-11-02','Aduu','Aduu','8.0','8.0','8.0','8.0','8.0','8.0','8.0','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-','-');"
	# sql=sql.strip(',')
	# sql=sql+");"
	# print(sql.replace("-","0"))